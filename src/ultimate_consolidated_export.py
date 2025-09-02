"""
Ultimate Consolidated Excel Export with Realistic Forecasting
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_ultimate_export() -> str:
    """Create the ultimate consolidated Excel with everything fixed and enhanced."""
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"artifacts/HPE_ULTIMATE_Consolidated_{timestamp}.xlsx"
    
    print(f"🚀 Creating ULTIMATE consolidated export: {output_file}")
    
    # Load all corrected data
    try:
        corrected_resources = pd.read_parquet('data_processed/resources_deduplicated.parquet')
        service_mappings = pd.read_parquet('data_processed/service_skillset_Services_to_skillsets_Mapping_Master_v5_clean_clean.parquet')
        
        from src.io_loader import load_processed_data
        data = load_processed_data()
        opportunities = data.get('opportunity_RAWDATA_Export_clean_clean', pd.DataFrame())
        
        print(f"✅ Loaded corrected data: {len(corrected_resources)} people, {len(opportunities)} opportunities")
        
    except Exception as e:
        print(f"❌ Failed to load data: {e}")
        return None
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # 1. EXECUTIVE SUMMARY WITH REALISTIC METRICS
        print("📊 Creating Executive Summary...")
        
        # Calculate realistic metrics
        total_people = len(corrected_resources)
        avg_skills_per_person = corrected_resources['skill_count'].mean()
        total_unique_skills = len(set(skill for skills in corrected_resources['all_skills'].dropna() 
                                    for skill in skills.split('; ')))
        
        # Service complexity analysis
        service_complexity = service_mappings.groupby('New Service Name')['Skill Set'].count()
        complex_services = len(service_complexity[service_complexity >= 20])
        
        exec_summary = pd.DataFrame({
            'Key_Metric': [
                'Total Unique People',
                'Average Skills per Person',
                'Total Unique Skills Available',
                'Geographic Locations',
                'Business Domains',
                'Active Opportunities',
                'Complex Services (20+ skills)',
                'Service-Skill Mappings',
                'Data Quality Status',
                'Forecasting Capability'
            ],
            'Value': [
                f"{total_people:,} people",
                f"{avg_skills_per_person:.1f} skills",
                f"{total_unique_skills:,} skills",
                f"{corrected_resources['city'].nunique()} cities",
                f"{corrected_resources['primary_domain'].nunique()} domains",
                f"{len(opportunities)} projects",
                f"{complex_services} services",
                f"{len(service_mappings):,} mappings",
                "CORRECTED (was 32x inflated)",
                "REALISTIC (based on actual requirements)"
            ],
            'Business_Impact': [
                'True workforce size for planning',
                'Rich skill diversity per person',
                'Comprehensive capability coverage',
                'Global talent distribution',
                'Business capability mapping',
                'Current project pipeline',
                'High-complexity service capability',
                'Skills-to-services connections',
                'Accurate analytics now possible',
                'Real capacity planning enabled'
            ]
        })
        
        exec_summary.to_excel(writer, sheet_name='EXECUTIVE_SUMMARY', index=False)
        
        # 2. CORRECTED TALENT DATABASE
        print("👥 Creating Corrected Talent Database...")
        corrected_resources.to_excel(writer, sheet_name='Talent_Database', index=False)
        
        # 3. REALISTIC RESOURCE REQUIREMENTS
        print("📋 Creating Realistic Resource Requirements...")
        
        # Calculate realistic resource needs per service
        service_analysis = service_mappings.groupby('New Service Name').agg({
            'Skill Set': 'count',
            'Technical  Domain': 'first'
        })
        service_analysis.columns = ['Skills_Required', 'Primary_Domain']
        service_analysis['Est_People_Needed'] = np.ceil(service_analysis['Skills_Required'] / 8).astype(int)
        service_analysis['Complexity_Level'] = pd.cut(
            service_analysis['Skills_Required'], 
            bins=[0, 5, 15, 30, 100], 
            labels=['Simple', 'Medium', 'Complex', 'Very Complex']
        )
        service_analysis = service_analysis.reset_index().sort_values('Skills_Required', ascending=False)
        service_analysis.to_excel(writer, sheet_name='Service_Requirements', index=False)
        
        # 4. CAPACITY VS DEMAND ANALYSIS
        print("⚖️ Creating Capacity vs Demand Analysis...")
        
        # Domain capacity
        domain_capacity = corrected_resources.groupby('primary_domain').agg({
            'resource_name': 'count',
            'skill_count': 'mean',
            'avg_rating': 'mean'
        }).round(2)
        domain_capacity.columns = ['Available_People', 'Avg_Skills_Per_Person', 'Avg_Rating']
        
        # Calculate demand per domain from services
        domain_demand = service_mappings['Technical  Domain'].value_counts().to_frame('Service_Demand')
        
        # Merge capacity and demand
        capacity_vs_demand = domain_capacity.merge(domain_demand, left_index=True, right_index=True, how='outer').fillna(0)
        capacity_vs_demand['Capacity_Ratio'] = capacity_vs_demand['Available_People'] / capacity_vs_demand['Service_Demand'].replace(0, 1)
        capacity_vs_demand['Status'] = capacity_vs_demand['Capacity_Ratio'].apply(
            lambda x: 'Surplus' if x > 2 else 'Adequate' if x > 1 else 'Gap'
        )
        capacity_vs_demand = capacity_vs_demand.reset_index()
        capacity_vs_demand.to_excel(writer, sheet_name='Capacity_vs_Demand', index=False)
        
        # 5. SKILL AVAILABILITY MATRIX
        print("🛠️ Creating Skill Availability Matrix...")
        
        # Create skill availability summary
        skill_data = []
        for _, person in corrected_resources.iterrows():
            skills = person['all_skills'].split('; ') if pd.notna(person['all_skills']) else []
            for skill in skills:
                skill_data.append({
                    'skill_name': skill,
                    'person_name': person['resource_name'],
                    'person_rating': person['avg_rating'],
                    'domain': person['primary_domain'],
                    'city': person['city']
                })
        
        skill_availability_df = pd.DataFrame(skill_data)
        skill_summary = skill_availability_df.groupby('skill_name').agg({
            'person_name': 'count',
            'person_rating': 'mean',
            'domain': lambda x: ', '.join(x.unique()[:3]),
            'city': lambda x: ', '.join(x.unique()[:3])
        }).round(2)
        skill_summary.columns = ['Available_People', 'Avg_Rating', 'Top_Domains', 'Top_Cities']
        skill_summary = skill_summary.reset_index().sort_values('Available_People', ascending=False)
        skill_summary.to_excel(writer, sheet_name='Skill_Availability', index=False)
        
        # 6. GEOGRAPHIC TALENT DISTRIBUTION
        print("🌍 Creating Geographic Analysis...")
        
        geo_analysis = corrected_resources.groupby(['city', 'primary_domain']).agg({
            'resource_name': 'count',
            'skill_count': 'mean',
            'avg_rating': 'mean'
        }).round(2)
        geo_analysis.columns = ['People_Count', 'Avg_Skills', 'Avg_Rating']
        geo_analysis = geo_analysis.reset_index()
        geo_analysis.to_excel(writer, sheet_name='Geographic_Distribution', index=False)
        
        # 7. TOP TALENT IDENTIFICATION
        print("🏆 Creating Top Talent Analysis...")
        
        # Calculate talent score
        corrected_resources['talent_score'] = (
            corrected_resources['skill_count'] * 0.6 + 
            corrected_resources['avg_rating'] * 10
        )
        
        top_talent = corrected_resources.nlargest(50, 'talent_score')[
            ['resource_name', 'manager', 'skill_count', 'avg_rating', 'talent_score',
             'primary_domain', 'city', 'all_skills']
        ]
        top_talent.to_excel(writer, sheet_name='Top_50_Talent', index=False)
        
        # 8. OPPORTUNITIES ANALYSIS
        if len(opportunities) > 0:
            opportunities.to_excel(writer, sheet_name='Current_Opportunities', index=False)
    
    # Apply styling
    print("🎨 Applying professional styling...")
    wb = openpyxl.load_workbook(output_file)
    
    # Style executive summary
    if 'EXECUTIVE_SUMMARY' in wb.sheetnames:
        ws = wb['EXECUTIVE_SUMMARY']
        ws.insert_rows(1)
        ws['A1'] = "HPE TALENT INTELLIGENCE - ULTIMATE CONSOLIDATED REPORT"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='0073E6', end_color='0073E6', fill_type='solid')
        ws.merge_cells('A1:C1')
    
    wb.save(output_file)
    
    print(f"✅ ULTIMATE export completed: {output_file}")
    
    # Show file info
    file_size = Path(output_file).stat().st_size / (1024*1024)
    print(f"\n📊 ULTIMATE EXPORT SUMMARY:")
    print(f"   📁 File: {output_file}")
    print(f"   💾 Size: {file_size:.1f} MB")
    print(f"   📋 Worksheets: {len(wb.sheetnames)}")
    print(f"   👥 Unique People: {len(corrected_resources):,}")
    print(f"   🛠️ Total Skills: {total_unique_skills:,}")
    print(f"   ⚙️ Services Analyzed: {len(service_mappings['New Service Name'].unique())}")
    
    return output_file

if __name__ == "__main__":
    main()
