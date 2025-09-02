"""
Enhanced Excel Export with all business intelligence enhancements.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_enhanced_excel_export() -> str:
    """Create comprehensive Excel export with all enhancements."""
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"artifacts/HPE_Enhanced_Intelligence_Export_{timestamp}.xlsx"
    
    print(f"🚀 Creating enhanced Excel export: {output_file}")
    
    # Load all data
    try:
        from src.io_loader import load_processed_data
        data = load_processed_data()
        
        # Load enhanced resource data
        enhanced_df = pd.read_parquet('data_processed/resource_DETAILS_28_Export_clean_clean_enhanced.parquet')
        
    except Exception as e:
        print(f"❌ Failed to load data: {e}")
        return None
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # 1. ENHANCED RESOURCE INTELLIGENCE
        print("📊 Creating Enhanced Resource Intelligence sheet...")
        enhanced_df.to_excel(writer, sheet_name='Enhanced_Resources', index=False)
        
        # 2. BUSINESS INTELLIGENCE SUMMARY
        print("📈 Creating Business Intelligence Summary...")
        
        # Skill category analysis
        skill_summary = enhanced_df.groupby(['Skill_Category', 'Experience_Category']).agg({
            'resource_name': 'count',
            'Resource_Value_Score': 'mean',
            'Seniority_Score': 'mean'
        }).round(2)
        skill_summary.columns = ['Resource_Count', 'Avg_Value_Score', 'Avg_Seniority']
        skill_summary = skill_summary.reset_index()
        skill_summary.to_excel(writer, sheet_name='Skill_Intelligence', index=False)
        
        # Geographic intelligence
        geo_summary = enhanced_df.groupby(['Region_Category', 'Domain_Category']).agg({
            'resource_name': 'count',
            'Resource_Value_Score': 'mean',
            'Skill_Complexity': 'mean'
        }).round(2)
        geo_summary.columns = ['Resource_Count', 'Avg_Value_Score', 'Avg_Complexity']
        geo_summary = geo_summary.reset_index()
        geo_summary.to_excel(writer, sheet_name='Geographic_Intelligence', index=False)
        
        # 3. HIGH-VALUE RESOURCES
        print("🏆 Creating High-Value Resources analysis...")
        top_resources = enhanced_df.nlargest(100, 'Resource_Value_Score')[
            ['resource_name', 'Skill_Category', 'Experience_Category', 'Region_Category', 
             'Domain_Category', 'Resource_Value_Score', 'Skill_Scarcity', 'manager']
        ]
        top_resources.to_excel(writer, sheet_name='Top_100_Resources', index=False)
        
        # 4. SCARCITY ANALYSIS
        print("💎 Creating Skill Scarcity Analysis...")
        scarcity_analysis = enhanced_df.groupby(['Skill_Scarcity', 'Skill_Category']).agg({
            'resource_name': 'count',
            'Resource_Value_Score': 'mean'
        }).round(2)
        scarcity_analysis.columns = ['Resource_Count', 'Avg_Value_Score']
        scarcity_analysis = scarcity_analysis.reset_index()
        scarcity_analysis.to_excel(writer, sheet_name='Scarcity_Analysis', index=False)
        
        # 5. EXECUTIVE DASHBOARD DATA
        print("📊 Creating Executive Dashboard data...")
        
        # Create executive summary
        exec_summary = {
            'Metric': [
                'Total Resources',
                'Skill Categories', 
                'Geographic Regions',
                'Domain Categories',
                'Expert-Level Resources',
                'High-Scarcity Skills',
                'Programming Specialists',
                'Cloud Specialists',
                'Security Specialists',
                'Average Resource Value Score'
            ],
            'Value': [
                f"{len(enhanced_df):,}",
                f"{enhanced_df['Skill_Category'].nunique()}",
                f"{enhanced_df['Region_Category'].nunique()}",
                f"{enhanced_df['Domain_Category'].nunique()}",
                f"{len(enhanced_df[enhanced_df['Experience_Category'] == 'Expert (5+ years)']):,}",
                f"{len(enhanced_df[enhanced_df['Skill_Scarcity'] == 'High Scarcity']):,}",
                f"{len(enhanced_df[enhanced_df['Skill_Category'] == 'Programming Languages']):,}",
                f"{len(enhanced_df[enhanced_df['Skill_Category'] == 'Cloud Platforms']):,}",
                f"{len(enhanced_df[enhanced_df['Skill_Category'] == 'Security & Compliance']):,}",
                f"{enhanced_df['Resource_Value_Score'].mean():.2f}"
            ]
        }
        
        exec_df = pd.DataFrame(exec_summary)
        exec_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
        
        # 6. ORIGINAL DATA SHEETS (for reference)
        if 'resource_DETAILS_28_Export_clean_clean' in data:
            data['resource_DETAILS_28_Export_clean_clean'].to_excel(writer, sheet_name='Original_Resources', index=False)
        
        if 'opportunity_RAWDATA_Export_clean_clean' in data:
            data['opportunity_RAWDATA_Export_clean_clean'].to_excel(writer, sheet_name='Original_Opportunities', index=False)
        
        if 'service_skillset_Services_to_skillsets_Mapping_Master_v5_clean_clean' in data:
            data['service_skillset_Services_to_skillsets_Mapping_Master_v5_clean_clean'].to_excel(writer, sheet_name='Service_Mappings', index=False)
    
    # Apply professional styling
    print("🎨 Applying professional styling...")
    wb = openpyxl.load_workbook(output_file)
    
    # Style key sheets
    for sheet_name in ['Enhanced_Resources', 'Skill_Intelligence', 'Executive_Summary']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Header styling
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    
    print(f"✅ Enhanced Excel export completed: {output_file}")
    
    # Show summary
    print(f"\n📊 EXPORT SUMMARY:")
    print(f"   📁 File: {output_file}")
    print(f"   📊 Worksheets: {len(wb.sheetnames)}")
    print(f"   📋 Enhanced Resources: {len(enhanced_df):,} records")
    print(f"   🆕 New Columns: 21 business intelligence columns")
    print(f"   💾 File Size: ~{Path(output_file).stat().st_size / (1024*1024):.1f} MB")
    
    return output_file

def main():
    """Main function to create enhanced export."""
    try:
        output_file = create_enhanced_excel_export()
        
        if output_file:
            print(f"\n🎉 SUCCESS! Enhanced Excel export ready:")
            print(f"   📁 {output_file}")
            print(f"\n🎯 What's New:")
            print(f"   ✅ 21 new business intelligence columns")
            print(f"   ✅ Skill categorization and scarcity analysis")
            print(f"   ✅ Geographic intelligence and timezone data")
            print(f"   ✅ Experience standardization and value scoring")
            print(f"   ✅ Executive summary with key metrics")
            print(f"   ✅ Professional styling and formatting")
    
    except Exception as e:
        print(f"❌ Export failed: {e}")

if __name__ == "__main__":
    main()
