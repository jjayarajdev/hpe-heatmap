"""
Comprehensive Excel Export Module for HPE Talent Intelligence Platform.

This module creates a unified Excel workbook with all data mapped on common columns,
providing business users with a complete view of skills, services, resources, and opportunities.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import json
from typing import Dict, List, Any, Optional
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from .utils import config, logger, Timer
    from .io_loader import load_processed_data
except ImportError:
    # For standalone execution
    import sys
    import os
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from utils import config, logger, Timer
    from io_loader import load_processed_data

class ExcelExporter:
    """Comprehensive Excel export functionality."""
    
    def __init__(self):
        self.data = {}
        self.taxonomy = {}
        self.export_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
    def load_all_data(self) -> bool:
        """Load all processed data and taxonomy."""
        try:
            # Load processed parquet files
            self.data = load_processed_data()
            
            # Load taxonomy if available
            taxonomy_path = Path("artifacts/taxonomy.json")
            if taxonomy_path.exists():
                with open(taxonomy_path, 'r') as f:
                    self.taxonomy = json.load(f)
            
            logger.info(f"Loaded {len(self.data)} datasets and taxonomy")
            return True
            
        except Exception as e:
            logger.error(f"Failed to load data: {e}")
            return False
    
    def create_unified_dataset(self) -> pd.DataFrame:
        """Create a unified dataset with all entities mapped on common columns."""
        logger.info("Creating unified dataset...")
        
        unified_records = []
        
        # Process Opportunities
        if 'opportunity_RAWDATA_Export_clean_clean' in self.data:
            opp_df = self.data['opportunity_RAWDATA_Export_clean_clean']
            for _, row in opp_df.iterrows():
                unified_records.append({
                    'Entity_Type': 'Opportunity',
                    'Entity_ID': row.get('opportunity_id', f"OPP_{len(unified_records)}"),
                    'Entity_Name': row.get('RR name', ''),
                    'Project_Name': row.get('RR Project Name', ''),
                    'Practice': row.get('RR Practice Name', ''),
                    'Delivery_Method': row.get('RR Delivery Method', ''),
                    'Location': row.get('RR Delivery Location', ''),
                    'Geography': row.get('RR Delivery Location Geo', ''),
                    'Status': row.get('RR Status', ''),
                    'Year_Created': row.get('RR_Created_Year', ''),
                    'Owner': row.get('RR Owner Name', ''),
                    'Assigned_Resource': row.get('RR Resource Assigned Name', ''),
                    'Resource_Email': row.get('RR Resource Assigned Email', ''),
                    'Required_Skill': row.get('Skill_Certification_Name', ''),
                    'Skill_Rating_Required': row.get('Minimum_Rating', ''),
                    'Skill_Rating_Actual': row.get('Resource Certification/Skill Rating', ''),
                    'Domain': row.get('RMR_Domain', ''),
                    'SubDomain': row.get('RMR_SubDomain', ''),
                    'MRU': row.get('RMR_MRU', ''),
                    'Compliance_Percentage': row.get('Compliance Percentage', 0),
                    'Skills_Match_Status': row.get('Matched Skills', ''),
                    'External_Resource': row.get('External Resource', ''),
                    'Service_Name': '',
                    'Skillset_Name': '',
                    'Manager': '',
                    'City': ''
                })
        
        # Process Resources
        if 'resource_DETAILS_28_Export_clean_clean' in self.data:
            res_df = self.data['resource_DETAILS_28_Export_clean_clean']
            for _, row in res_df.iterrows():
                unified_records.append({
                    'Entity_Type': 'Resource',
                    'Entity_ID': row.get('resource_id', f"RES_{len(unified_records)}"),
                    'Entity_Name': row.get('resource_name', ''),
                    'Project_Name': '',
                    'Practice': '',
                    'Delivery_Method': '',
                    'Location': '',
                    'Geography': '',
                    'Status': 'Active',
                    'Year_Created': '',
                    'Owner': '',
                    'Assigned_Resource': row.get('resource_name', ''),
                    'Resource_Email': '',
                    'Required_Skill': '',
                    'Skill_Rating_Required': '',
                    'Skill_Rating_Actual': row.get('Rating', ''),
                    'Domain': row.get('domain', ''),
                    'SubDomain': '',
                    'MRU': row.get('RMR_MRU', ''),
                    'Compliance_Percentage': 0,
                    'Skills_Match_Status': '',
                    'External_Resource': 'Internal',
                    'Service_Name': '',
                    'Skillset_Name': row.get('Skill_Set_Name', ''),
                    'Manager': row.get('manager', ''),
                    'City': row.get('RMR_City', ''),
                    'Skill_Name': row.get('Skill_Certification_Name', ''),
                    'Proficiency_Rating': row.get('Proficieny_Rating', '')
                })
        
        # Process Service-Skillset Mappings
        if 'service_skillset_Services_to_skillsets_Mapping_Master_v5_clean_clean' in self.data:
            svc_df = self.data['service_skillset_Services_to_skillsets_Mapping_Master_v5_clean_clean']
            for _, row in svc_df.iterrows():
                unified_records.append({
                    'Entity_Type': 'Service_Skillset_Mapping',
                    'Entity_ID': f"MAP_{len(unified_records)}",
                    'Entity_Name': 'Service-Skillset Mapping',
                    'Project_Name': '',
                    'Practice': '',
                    'Delivery_Method': '',
                    'Location': '',
                    'Geography': '',
                    'Status': 'Active',
                    'Year_Created': '',
                    'Owner': '',
                    'Assigned_Resource': '',
                    'Resource_Email': '',
                    'Required_Skill': '',
                    'Skill_Rating_Required': '',
                    'Skill_Rating_Actual': '',
                    'Domain': row.get('Technical  Domain', ''),
                    'SubDomain': '',
                    'MRU': '',
                    'Compliance_Percentage': 0,
                    'Skills_Match_Status': '',
                    'External_Resource': '',
                    'Service_Name': row.get('New Service Name', ''),
                    'Skillset_Name': row.get('Skill Set', ''),
                    'Manager': '',
                    'City': '',
                    'FY25_PL': row.get('FY25 PL', ''),
                    'FY25_PL_Name': row.get('FY25 PL Name', ''),
                    'FY25_Focus_Area': row.get('FY25 Focus Area', ''),
                    'Mandatory_Optional': row.get('Mandatory/ Optional', ''),
                    'CYB_Alignment': row.get('CYB alignment to other domains', '')
                })
        
        # Process Skills from taxonomy
        if 'skills' in self.taxonomy:
            for skill in self.taxonomy['skills']:
                unified_records.append({
                    'Entity_Type': 'Skill',
                    'Entity_ID': skill.get('id', ''),
                    'Entity_Name': skill.get('name', ''),
                    'Project_Name': '',
                    'Practice': '',
                    'Delivery_Method': '',
                    'Location': '',
                    'Geography': '',
                    'Status': 'Active',
                    'Year_Created': '',
                    'Owner': '',
                    'Assigned_Resource': '',
                    'Resource_Email': '',
                    'Required_Skill': '',
                    'Skill_Rating_Required': '',
                    'Skill_Rating_Actual': '',
                    'Domain': skill.get('category', ''),
                    'SubDomain': '',
                    'MRU': '',
                    'Compliance_Percentage': 0,
                    'Skills_Match_Status': '',
                    'External_Resource': '',
                    'Service_Name': '',
                    'Skillset_Name': '',
                    'Manager': '',
                    'City': '',
                    'Skill_Name': skill.get('name', ''),
                    'Skill_Description': skill.get('description', ''),
                    'Skill_Hierarchy': skill.get('hierarchy', ''),
                    'Skill_Synonyms': ', '.join(skill.get('synonyms', []))
                })
        
        # Process Services from taxonomy
        if 'services' in self.taxonomy:
            for service in self.taxonomy['services']:
                unified_records.append({
                    'Entity_Type': 'Service',
                    'Entity_ID': service.get('id', ''),
                    'Entity_Name': service.get('name', ''),
                    'Project_Name': '',
                    'Practice': '',
                    'Delivery_Method': '',
                    'Location': '',
                    'Geography': '',
                    'Status': 'Active',
                    'Year_Created': '',
                    'Owner': '',
                    'Assigned_Resource': '',
                    'Resource_Email': '',
                    'Required_Skill': '',
                    'Skill_Rating_Required': '',
                    'Skill_Rating_Actual': '',
                    'Domain': service.get('category', ''),
                    'SubDomain': '',
                    'MRU': '',
                    'Compliance_Percentage': 0,
                    'Skills_Match_Status': '',
                    'External_Resource': '',
                    'Service_Name': service.get('name', ''),
                    'Skillset_Name': '',
                    'Manager': '',
                    'City': '',
                    'Service_Description': service.get('description', ''),
                    'Service_Aliases': ', '.join(service.get('aliases', []))
                })
        
        # Process Skillsets from taxonomy
        if 'skillsets' in self.taxonomy:
            for skillset in self.taxonomy['skillsets']:
                unified_records.append({
                    'Entity_Type': 'Skillset',
                    'Entity_ID': skillset.get('id', ''),
                    'Entity_Name': skillset.get('name', ''),
                    'Project_Name': '',
                    'Practice': '',
                    'Delivery_Method': '',
                    'Location': '',
                    'Geography': '',
                    'Status': 'Active',
                    'Year_Created': '',
                    'Owner': '',
                    'Assigned_Resource': '',
                    'Resource_Email': '',
                    'Required_Skill': '',
                    'Skill_Rating_Required': '',
                    'Skill_Rating_Actual': '',
                    'Domain': skillset.get('category', ''),
                    'SubDomain': '',
                    'MRU': '',
                    'Compliance_Percentage': 0,
                    'Skills_Match_Status': '',
                    'External_Resource': '',
                    'Service_Name': '',
                    'Skillset_Name': skillset.get('name', ''),
                    'Manager': '',
                    'City': '',
                    'Skillset_Description': skillset.get('description', ''),
                    'Skillset_Hierarchy': skillset.get('hierarchy', '')
                })
        
        unified_df = pd.DataFrame(unified_records)
        logger.info(f"Created unified dataset with {len(unified_df)} records")
        return unified_df
    
    def create_summary_analytics(self) -> Dict[str, pd.DataFrame]:
        """Create summary analytics for business insights."""
        summaries = {}
        
        # Skills by Domain Summary
        if 'resource_DETAILS_28_Export_clean_clean' in self.data:
            res_df = self.data['resource_DETAILS_28_Export_clean_clean']
            
            # Convert rating to numeric for aggregation
            res_df_copy = res_df.copy()
            res_df_copy['Rating_Numeric'] = pd.to_numeric(
                res_df_copy['Rating'].astype(str).str.extract('(\d+)', expand=False), 
                errors='coerce'
            )
            
            skills_summary = res_df_copy.groupby(['domain', 'Skill_Certification_Name']).agg({
                'resource_name': 'count',
                'Rating_Numeric': ['mean', 'min', 'max'],
                'RMR_MRU': lambda x: ', '.join(x.dropna().unique()[:5])
            }).round(2)
            
            skills_summary.columns = ['Resource_Count', 'Avg_Rating', 'Min_Rating', 'Max_Rating', 'Top_MRUs']
            skills_summary = skills_summary.reset_index()
            summaries['Skills_by_Domain'] = skills_summary
        
        # Services by Focus Area Summary  
        if 'service_skillset_Services_to_skillsets_Mapping_Master_v5_clean_clean' in self.data:
            svc_df = self.data['service_skillset_Services_to_skillsets_Mapping_Master_v5_clean_clean']
            
            services_summary = svc_df.groupby(['FY25 Focus Area', 'Technical  Domain']).agg({
                'New Service Name': 'count',
                'Skill Set': 'nunique',
                'Mandatory/ Optional': lambda x: (x == 'Mandatory').sum()
            })
            
            services_summary.columns = ['Service_Count', 'Unique_Skillsets', 'Mandatory_Skills']
            services_summary = services_summary.reset_index()
            summaries['Services_by_Focus_Area'] = services_summary
        
        # Resource Distribution Summary
        if 'resource_DETAILS_28_Export_clean_clean' in self.data:
            res_df = self.data['resource_DETAILS_28_Export_clean_clean']
            
            # Convert rating to numeric for aggregation
            res_df_copy = res_df.copy()
            res_df_copy['Rating_Numeric'] = pd.to_numeric(
                res_df_copy['Rating'].astype(str).str.extract('(\d+)', expand=False), 
                errors='coerce'
            )
            
            resource_summary = res_df_copy.groupby(['RMR_MRU', 'domain']).agg({
                'resource_name': 'nunique',
                'Rating_Numeric': 'mean',
                'RMR_City': lambda x: ', '.join(x.dropna().unique()[:3])
            }).round(2)
            
            resource_summary.columns = ['Unique_Resources', 'Avg_Rating', 'Top_Cities']
            resource_summary = resource_summary.reset_index()
            summaries['Resource_Distribution'] = resource_summary
        
        # Opportunity Analysis
        if 'opportunity_RAWDATA_Export_clean_clean' in self.data:
            opp_df = self.data['opportunity_RAWDATA_Export_clean_clean']
            
            # Convert compliance percentage to numeric
            opp_df_copy = opp_df.copy()
            opp_df_copy['Compliance Percentage'] = pd.to_numeric(
                opp_df_copy['Compliance Percentage'], errors='coerce'
            )
            
            opp_summary = opp_df_copy.groupby(['RR Practice Name', 'RR Status']).agg({
                'RR name': 'count',
                'Compliance Percentage': 'mean',
                'Matched Skills': lambda x: (x.astype(str) == 'Match').sum()
            }).round(2)
            
            opp_summary.columns = ['Opportunity_Count', 'Avg_Compliance', 'Matched_Count']
            opp_summary = opp_summary.reset_index()
            summaries['Opportunity_Analysis'] = opp_summary
        
        return summaries
    
    def style_worksheet(self, worksheet, df: pd.DataFrame, title: str):
        """Apply professional styling to worksheet."""
        # Set title
        worksheet.insert_rows(1, 2)
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        worksheet['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        worksheet.merge_cells('A1:' + chr(ord('A') + len(df.columns) - 1) + '1')
        
        # Add export timestamp
        worksheet['A2'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        worksheet['A2'].font = Font(size=10, italic=True)
        
        # Style headers
        header_row = 3
        for col_idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.value = column
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), header_row + 1):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                
                # Alternate row colors
                if r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=header_row, 
                                     max_row=len(df) + header_row,
                                     min_col=1, 
                                     max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
    
    def export_to_excel(self, output_path: str = None) -> str:
        """Export all data to a comprehensive Excel file."""
        if not self.load_all_data():
            raise Exception("Failed to load data")
        
        if output_path is None:
            output_path = f"artifacts/HPE_Talent_Intelligence_Export_{self.export_timestamp}.xlsx"
        
        logger.info(f"Creating Excel export: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # 1. Unified Dataset
            unified_df = self.create_unified_dataset()
            unified_df.to_excel(writer, sheet_name='Unified_Data', index=False)
            
            # 2. Summary Analytics
            summaries = self.create_summary_analytics()
            for sheet_name, summary_df in summaries.items():
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 3. Raw Data Sheets
            for data_name, df in self.data.items():
                # Truncate sheet name to Excel limit
                sheet_name = data_name[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 4. Taxonomy Sheets
            if self.taxonomy:
                if 'services' in self.taxonomy:
                    services_df = pd.DataFrame(self.taxonomy['services'])
                    services_df.to_excel(writer, sheet_name='Services_Taxonomy', index=False)
                
                if 'skillsets' in self.taxonomy:
                    skillsets_df = pd.DataFrame(self.taxonomy['skillsets'])
                    skillsets_df.to_excel(writer, sheet_name='Skillsets_Taxonomy', index=False)
                
                if 'skills' in self.taxonomy:
                    skills_df = pd.DataFrame(self.taxonomy['skills'])
                    skills_df.to_excel(writer, sheet_name='Skills_Taxonomy', index=False)
        
        # Apply styling
        wb = openpyxl.load_workbook(output_path)
        
        # Style key sheets
        key_sheets = {
            'Unified_Data': 'HPE Talent Intelligence - Unified Dataset',
            'Skills_by_Domain': 'Skills Analysis by Domain',
            'Services_by_Focus_Area': 'Services Analysis by Focus Area',
            'Resource_Distribution': 'Resource Distribution Analysis',
            'Opportunity_Analysis': 'Opportunity Analysis'
        }
        
        for sheet_name, title in key_sheets.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                df = unified_df if sheet_name == 'Unified_Data' else summaries.get(sheet_name, pd.DataFrame())
                if not df.empty:
                    # Clear existing content
                    ws.delete_rows(1, ws.max_row)
                    # Apply styling
                    self.style_worksheet(ws, df, title)
        
        wb.save(output_path)
        logger.info(f"Excel export completed: {output_path}")
        
        return output_path

def create_comprehensive_export() -> str:
    """Main function to create comprehensive Excel export."""
    exporter = ExcelExporter()
    return exporter.export_to_excel()

def main():
    """CLI entry point."""
    try:
        output_file = create_comprehensive_export()
        print(f"✅ Excel export created successfully: {output_file}")
        
        # Print summary
        exporter = ExcelExporter()
        exporter.load_all_data()
        unified_df = exporter.create_unified_dataset()
        
        print(f"\n📊 Export Summary:")
        print(f"   Total Records: {len(unified_df):,}")
        print(f"   Entity Types: {unified_df['Entity_Type'].nunique()}")
        print(f"   Unique Domains: {unified_df['Domain'].nunique()}")
        print(f"   Unique MRUs: {unified_df['MRU'].nunique()}")
        
        entity_counts = unified_df['Entity_Type'].value_counts()
        for entity_type, count in entity_counts.items():
            print(f"   {entity_type}: {count:,} records")
        
    except Exception as e:
        logger.error(f"Export failed: {e}")
        print(f"❌ Export failed: {e}")

if __name__ == "__main__":
    main()
